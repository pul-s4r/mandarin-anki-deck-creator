from __future__ import annotations

import csv
import io
from datetime import UTC, datetime
from pathlib import Path

import pytest

from anki_deck_generator.config.source_sets import (
    CsvExporterConfig,
    LocalFileSource,
    SourceSet,
    XlsxExporterConfig,
    load_source_sets_yaml,
)
from anki_deck_generator.dictionary.enrich import VocabularyRow
from anki_deck_generator.errors import IntegrationError
from anki_deck_generator.export.csv_writer import FIELDNAMES, vocabulary_csv_bytes
from anki_deck_generator.export.exporter_factory import (
    build_file_exporters_from_configs,
    resolve_exporters_for_schedule,
)
from anki_deck_generator.export.exporters import (
    VocabularyCsvFileExporter,
    VocabularyXlsxExporter,
    VocabularyXlsxFileExporter,
)
from anki_deck_generator.export.xlsx_writer import METADATA_SHEET, VOCABULARY_SHEET, vocabulary_xlsx_bytes
from anki_deck_generator.pipeline_types import PipelineResult, PipelineStats
from anki_deck_generator.sync.report import SyncReport, SyncReportStats


def _sample_rows() -> list[VocabularyRow]:
    return [
        VocabularyRow(
            key=1,
            simplified="的",
            traditional="的",
            pinyin="de",
            meaning="particle",
            part_of_speech="particle",
            usage_notes="",
            sentence_simplified="好的",
            sentence_traditional="好的",
            sentence_pinyin="hǎo de",
            sentence_meaning="okay",
        )
    ]


def _empty_pipeline_result(rows: list[VocabularyRow]) -> PipelineResult:
    return PipelineResult(
        rows=rows,
        sentence_links=[],
        stats=PipelineStats(
            block_count=0,
            chunk_count=1,
            raw_card_count=len(rows),
            deduped_card_count=len(rows),
            enriched_count=0,
            llm_translation_fallback_count=0,
            decomposition_fallback_count=0,
            sentence_link_count=0,
        ),
    )


def _csv_rows_from_bytes(data: bytes) -> list[dict[str, str]]:
    text = data.decode("utf-8")
    if text.startswith("\ufeff"):
        text = text[1:]
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _norm_xlsx_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _xlsx_vocab_rows(data: bytes) -> list[list[object]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb[VOCABULARY_SHEET]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _xlsx_metadata_map(data: bytes) -> dict[str, str]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb[METADATA_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Field", "Value")
    return {str(k): str(v) for k, v in rows[1:] if k is not None}


def test_xlsx_sheet_names_and_row_count() -> None:
    rows = _sample_rows()
    data = vocabulary_xlsx_bytes(rows)
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True)
    assert wb.sheetnames == [VOCABULARY_SHEET, METADATA_SHEET]
    ws = wb[VOCABULARY_SHEET]
    sheet_rows = list(ws.iter_rows(values_only=True))
    assert len(sheet_rows) == 1 + len(rows)


def test_xlsx_vocabulary_matches_csv_rows() -> None:
    rows = _sample_rows()
    csv_data = vocabulary_csv_bytes(rows, bom=False)
    xlsx_data = vocabulary_xlsx_bytes(rows)

    csv_rows = _csv_rows_from_bytes(csv_data)
    xlsx_table = _xlsx_vocab_rows(xlsx_data)
    assert xlsx_table[0] == list(FIELDNAMES)
    for i, csv_row in enumerate(csv_rows, start=1):
        xlsx_row = xlsx_table[i]
        for j, field in enumerate(FIELDNAMES):
            assert _norm_xlsx_cell(xlsx_row[j]) == csv_row[field]


def test_xlsx_metadata_contains_sync_report_fields() -> None:
    started = datetime(2024, 6, 1, 12, 0, 0, tzinfo=UTC)
    finished = datetime(2024, 6, 1, 12, 5, 0, tzinfo=UTC)
    report = SyncReport(
        run_id="run-123",
        run_started_at=started,
        run_finished_at=finished,
        stats=SyncReportStats(chunks_processed=3, chunks_skipped=1, documents_skipped=0, sources_processed=1),
        export_paths=["/tmp/deck.csv"],
        dry_run=False,
    )
    data = vocabulary_xlsx_bytes(_sample_rows(), sync_report=report, source_set_name="lessons")
    meta = _xlsx_metadata_map(data)
    assert meta["run_id"] == "run-123"
    assert meta["source_set"] == "lessons"
    assert meta["run_started_at"] == started.isoformat()
    assert meta["run_finished_at"] == finished.isoformat()
    assert meta["chunks_processed"] == "3"
    assert meta["chunks_skipped"] == "1"
    assert meta["documents_skipped"] == "0"
    assert meta["sources_processed"] == "1"
    assert "/tmp/deck.csv" in meta["export_paths"]
    assert meta["exporter_status"] == "none"


def test_xlsx_exporter_class_wraps_bytes() -> None:
    rows = _sample_rows()
    pr = _empty_pipeline_result(rows)
    exp = VocabularyXlsxExporter(source_set_name="t")
    assert exp.filename_suggestion == "vocabulary.xlsx"
    data = exp.export(pr)
    assert data.startswith(b"PK")


def test_missing_openpyxl_raises_integration_error(monkeypatch: pytest.MonkeyPatch) -> None:
    import builtins

    real_import = builtins.__import__

    def blocked_import(
        name: str,
        globals: object = None,
        locals: object = None,
        fromlist: object = (),
        level: int = 0,
    ) -> object:
        if name == "openpyxl" or name.startswith("openpyxl."):
            raise ImportError("blocked")
        return real_import(name, globals, locals, fromlist, level)

    monkeypatch.setattr(builtins, "__import__", blocked_import)
    with pytest.raises(IntegrationError, match="openpyxl"):
        vocabulary_xlsx_bytes(_sample_rows())


def test_build_file_exporters_from_source_set_configs(tmp_path: Path) -> None:
    csv_path = tmp_path / "deck.csv"
    xlsx_path = tmp_path / "deck.xlsx"
    configs = (
        CsvExporterConfig(type="csv", destination=csv_path),
        XlsxExporterConfig(type="xlsx", destination=xlsx_path),
    )
    exporters = build_file_exporters_from_configs(configs, csv_bom=True)
    assert len(exporters) == 2
    assert isinstance(exporters[0], VocabularyCsvFileExporter)
    assert isinstance(exporters[1], VocabularyXlsxFileExporter)
    assert exporters[0].output_path == csv_path.resolve()
    assert exporters[1].output_path == xlsx_path.resolve()


def test_load_source_set_yaml_exporters(tmp_path: Path) -> None:
    fixture = tmp_path / "sample.md"
    fixture.write_text("# hi", encoding="utf-8")
    csv_out = tmp_path / "out.csv"
    xlsx_out = tmp_path / "out.xlsx"
    yml = tmp_path / "sources.yaml"
    yml.write_text(
        f"""
schema_version: 1
source_sets:
  lessons:
    sources:
      - provider: local-filesystem
        path: {fixture}
    exporters:
      - type: csv
        destination: {csv_out}
      - type: xlsx
        destination: {xlsx_out}
""",
        encoding="utf-8",
    )
    cfg = load_source_sets_yaml(yml)
    ss = cfg["lessons"]
    assert len(ss.exporters) == 2
    assert isinstance(ss.exporters[0], CsvExporterConfig)
    assert isinstance(ss.exporters[1], XlsxExporterConfig)


def test_resolve_exporters_prefers_yaml_over_cli(tmp_path: Path) -> None:
    md = tmp_path / "n.md"
    md.write_text("x", encoding="utf-8")
    yaml_csv = tmp_path / "yaml.csv"
    yaml_xlsx = tmp_path / "yaml.xlsx"
    cli_csv = tmp_path / "cli.csv"
    sset = SourceSet(
        name="t",
        sources=(LocalFileSource(provider="local-filesystem", path=md, external_id="e"),),
        exporters=(
            CsvExporterConfig(type="csv", destination=yaml_csv),
            XlsxExporterConfig(type="xlsx", destination=yaml_xlsx),
        ),
    )
    exporters = resolve_exporters_for_schedule(sset, cli_output=cli_csv, csv_bom=False)
    paths = {exp.output_path for exp in exporters}
    assert yaml_csv.resolve() in paths
    assert yaml_xlsx.resolve() in paths
    assert cli_csv.resolve() not in paths


def test_resolve_exporters_requires_cli_when_yaml_missing_exporters(tmp_path: Path) -> None:
    md = tmp_path / "n.md"
    md.write_text("x", encoding="utf-8")
    sset = SourceSet(
        name="t",
        sources=(LocalFileSource(provider="local-filesystem", path=md, external_id="e"),),
    )
    with pytest.raises(ValueError, match="Configure exporters"):
        resolve_exporters_for_schedule(sset, cli_output=None)


def test_incremental_sync_writes_csv_and_xlsx_side_by_side(tmp_path: Path) -> None:
    from anki_deck_generator.config.settings import Settings
    from anki_deck_generator.export.exporter_factory import build_schedule_exporters_from_configs
    from anki_deck_generator.state.sqlite_store import SqliteStateStore
    from anki_deck_generator.sync.orchestrator import run_incremental_sync

    root = Path(__file__).resolve().parents[1] / "tests" / "baselines"
    md_path = root / "inputs" / "sample.md"
    settings = Settings(
        llm_fixture_path=root / "llm_mock.json",
        cedict_path=root / "cedict_sample.u8",
        enable_sentences=False,
        skip_lines_filter=False,
    )
    db = tmp_path / "state.db"
    store = SqliteStateStore(db)
    store.init_schema()
    ext_id = str(md_path.resolve())
    csv_out = tmp_path / "deck.csv"
    xlsx_out = tmp_path / "deck.xlsx"
    sset = SourceSet(
        name="lessons",
        sources=(LocalFileSource(provider="local-filesystem", path=md_path, external_id=ext_id),),
        exporters=(
            CsvExporterConfig(type="csv", destination=csv_out),
            XlsxExporterConfig(type="xlsx", destination=xlsx_out),
        ),
    )
    exporters = build_schedule_exporters_from_configs(sset.exporters, csv_bom=False)
    report = run_incremental_sync(
        sset,
        settings=settings,
        state_store=store,
        exporters=exporters,
    )
    assert csv_out.is_file()
    assert xlsx_out.is_file()
    assert str(csv_out) in report.export_paths
    assert str(xlsx_out) in report.export_paths

    csv_rows = _csv_rows_from_bytes(csv_out.read_bytes())
    xlsx_table = _xlsx_vocab_rows(xlsx_out.read_bytes())
    assert len(csv_rows) == len(xlsx_table) - 1
    for i, csv_row in enumerate(csv_rows, start=1):
        for j, field in enumerate(FIELDNAMES):
            assert _norm_xlsx_cell(xlsx_table[i][j]) == csv_row[field]

    meta = _xlsx_metadata_map(xlsx_out.read_bytes())
    assert meta["run_id"] == report.run_id
    assert meta["source_set"] == "lessons"

