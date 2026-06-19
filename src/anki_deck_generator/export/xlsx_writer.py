"""XLSX vocabulary export with optional run metadata worksheet."""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

from anki_deck_generator.dictionary.enrich import VocabularyRow
from anki_deck_generator.errors import IntegrationError
from anki_deck_generator.export.csv_writer import FIELDNAMES

if TYPE_CHECKING:
    from anki_deck_generator.sync.report import SyncReport

VOCABULARY_SHEET = "Vocabulary"
METADATA_SHEET = "Run metadata"


def _require_openpyxl() -> tuple[object, type]:
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - exercised via monkeypatch
        raise IntegrationError(
            "XLSX export requires openpyxl. Install with: pip install 'anki-deck-generator[xlsx]'"
        ) from exc
    return openpyxl, Workbook


def _row_values(row: VocabularyRow) -> list[object]:
    def cell(value: object) -> object:
        return "" if value is None else value

    return [
        cell(row.key),
        cell(row.simplified),
        cell(row.traditional),
        cell(row.pinyin),
        cell(row.meaning),
        cell(row.part_of_speech),
        cell(row.usage_notes),
        cell(row.sentence_simplified),
        cell(row.sentence_traditional),
        cell(row.sentence_pinyin),
        cell(row.sentence_meaning),
    ]


def sync_report_metadata_rows(
    sync_report: SyncReport | None,
    *,
    source_set_name: str | None = None,
) -> list[tuple[str, str]]:
    """Key-value rows for the Run metadata worksheet."""
    if sync_report is None:
        return [("source_set", source_set_name or ""), ("note", "No sync report attached")]

    rows: list[tuple[str, str]] = [
        ("run_id", sync_report.run_id),
        ("source_set", source_set_name or ""),
        ("run_started_at", sync_report.run_started_at.isoformat()),
        (
            "run_finished_at",
            sync_report.run_finished_at.isoformat() if sync_report.run_finished_at else "",
        ),
        ("dry_run", str(sync_report.dry_run)),
        ("chunks_processed", str(sync_report.stats.chunks_processed)),
        ("chunks_skipped", str(sync_report.stats.chunks_skipped)),
        ("documents_skipped", str(sync_report.stats.documents_skipped)),
        ("sources_processed", str(sync_report.stats.sources_processed)),
        ("outcomes_count", str(len(sync_report.outcomes))),
        ("export_paths", "; ".join(sync_report.export_paths)),
    ]
    if sync_report.exports:
        rows.append(("exporter_status", "present"))
    else:
        rows.append(("exporter_status", "none"))
    return rows


def vocabulary_xlsx_bytes(
    rows: list[VocabularyRow],
    *,
    sync_report: SyncReport | None = None,
    source_set_name: str | None = None,
) -> bytes:
    """Build an XLSX workbook: Vocabulary sheet plus optional Run metadata."""
    _, Workbook = _require_openpyxl()
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    ws_vocab = wb.create_sheet(VOCABULARY_SHEET)
    ws_vocab.append(list(FIELDNAMES))
    for row in rows:
        ws_vocab.append(_row_values(row))

    ws_meta = wb.create_sheet(METADATA_SHEET)
    ws_meta.append(["Field", "Value"])
    for field, value in sync_report_metadata_rows(sync_report, source_set_name=source_set_name):
        ws_meta.append([field, value])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
